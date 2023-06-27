import openpyxl
import requests
import xlwt
import win32com.client as win32
'''Updates data and downloads data files for all zones in the file'''
class LoaderUPS:
    '''Contains functions for updating data in the Carriers zone range'''
    def CreateEmptyXlsFile(path):
        '''Creates empty xls file.
        Arquments: path - path of the file
        returns: None'''
        wb=xlwt.Workbook()
        wb.add_sheet('1')
        wb.save(path)
        return wb
        
    def DownloadXlsFile(zipzone,path,workbook):
        '''Writes .xls file from the server content in existing file
        Arquments: zipzone = 3 first numbers of the zipcode. path - path of the file. workbook - file's workbook
        returns: None'''
        headers = { 'User-Agent': 'Chrome/92.0.4515.107'}
        resp=requests.get(url+zipzone+'.xls',headers=headers)
        open(path,'wb').write(resp.content)
        workbook.save(path)
    
    def ConvertXlsToXlsx(uploadpath,downloadpath):
        '''Converts .xls file to .xlxs file 
        Arquments: uploadpath - path of the .xls file. downloadpath - path and name where to save .xlsx file
        returns: None'''
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        wb = excel.Workbooks.Open(uploadpath)
        wb.SaveAs(downloadpath, FileFormat = 51)    #FileFormat = 51 is for .xlsx
        wb.Close()                               
        excel.Application.Quit()
    
    def UpdateZip(ZipCodeNumber,worksheet):
        '''copies zip codes of downloaded zones file to cell in Carriers zone range
         Arquments: ZipCodeNumber - number of row in Carriers zone range. worksheet - worksheet in Carriers zone range
         returns: None'''
        worksheet.cell(row = ZipCodeNumber, column = 2).value=s[39:42]+s[43:45]
        worksheet.cell(row = ZipcodeNumber, column = 3).value=s[49:52]+s[53:55]

url = 'https://www.ups.com/media/us/currentrates/zone-csv/'
wb=openpyxl.load_workbook("Carriers zone range.xlsx")
ws = wb["UPS zip ranges"]
for ZipCodeNumber in range(2,904):
    zipzone=ws.cell(row = ZipCodeNumber, column = 2).value[:3]#value from Carriers zone range
    path=f'D:/UPS_Carriers_zone_range_updater/xls_files/{zipzone}.xls'#Path where file should be saved
    wb=LoaderUPS.CreateEmptyXlsFile(path)
    LoaderUPS.DownloadXlsFile(zipzone=zipzone,path=path,workbook=wb)
    LoaderUPS.ConvertXlsToXlsx(uploadpath=f'xls_files/{zipzone}.xls',downloadpath=f'xlsx_files/{zipzone}.xlsx')
    LoaderUPS.UpdateZip(ZipcodeNumber=ZipcodeNumber,worksheet=ws)
wb.save("Carriers zone range.xlsx")
