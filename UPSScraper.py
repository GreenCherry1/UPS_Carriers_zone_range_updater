import openpyxl
import requests
import xlwt

#creates empty xls file
def CreateEmptyXlsFile(path):
    wb=xlwt.Workbook()
    wb.add_sheet('1')
    wb.save(path)
    return wb
    
#writes .xls file from the server content in existing file
#zipzone = 3 first numbers of the zipcode
def DownloadXlsFile(zipzone,path,workbook):
    headers = { 'User-Agent': 'Chrome/92.0.4515.107'}
    resp=requests.get(url+zipzone+'.xls',headers=headers)
    open(path,'wb').write(resp.content)
    workbook.save(path)

url = 'https://www.ups.com/media/us/currentrates/zone-csv/'
ws = openpyxl.load_workbook("Carriers zone range.xlsx")["UPS zip ranges"]
for ZipCodeNumber in range(2,904):
    zipzone=ws.cell(row = ZipCodeNumber, column = 2).value[:3]#value from Carriers zone range
    path=f'D:/UPS_Carriers_zone_range_updater/xls_files/{zipzone}.xls'#Path where file needs to be saved
    wb=CreateEmptyXlsFile(path)
    DownloadXlsFile(zipzone=zipzone,path=path,workbook=wb)
