import openpyxl
wb=openpyxl.load_workbook("Carriers zone range.xlsx")
ws = wb["UPS zip ranges"]
for i in range(2,904):
    s=openpyxl.load_workbook(f'xlsx_files/{ws.cell(row = i, column = 2).value[:3]}.xlsx').active.cell(row=5,column=1).value
    ws.cell(row = i, column = 2).value=s[39:42]+s[43:45]
    ws.cell(row = i, column = 3).value=s[49:52]+s[53:55]
wb.save("Carriers zone range.xlsx")
